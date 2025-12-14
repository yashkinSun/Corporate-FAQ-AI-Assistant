# doc_parser.py
import os
import re
import PyPDF2
import docx
import openpyxl

def parse_pdf(file_path: str) -> str:
    """Парсит PDF-файл и возвращает текст."""
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        # DEBUG: Проверим, сколько символов извлекли
        print(f"[DEBUG] parse_pdf: Read {len(text)} chars from {file_path}")
    except Exception as e:
        print(f"[ERROR] parse_pdf: Error parsing {file_path} - {e}")
    return text

def parse_docx(file_path: str) -> str:
    """Парсит DOCX-файл и возвращает текст."""
    text = ""
    try:
        doc = docx.Document(file_path)
        text = "\n".join(para.text for para in doc.paragraphs)
        # DEBUG
        print(f"[DEBUG] parse_docx: Read {len(text)} chars from {file_path}")
    except Exception as e:
        print(f"[ERROR] parse_docx: Error parsing {file_path} - {e}")
    return text

def parse_excel(file_path: str) -> str:
    all_text = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) for cell in row if cell is not None]
                if row_text:
                    all_text.append(" ".join(row_text))

        content = "\n".join(all_text)
        # DEBUG
        print(f"[DEBUG] parse_excel: Read {len(content)} chars from {file_path}")
        print(f"[DEBUG] parse_excel FULL TEXT from {file_path}:\n{content}")
        return content
    except Exception as e:
        print(f"[ERROR] parse_excel: Error parsing {file_path} - {e}")
        return ""


def _strip_html_tags(text: str) -> str:
    """Удаляет HTML-теги из текста."""
    return re.sub(r"<[^>]+>", "", text)


def _normalize_links(text: str) -> str:
    """Преобразует markdown-ссылки в формат plain text."""
    return re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1 (\2)", text)


def parse_markdown(file_path: str) -> str:
    """Парсит Markdown-файл и возвращает нормализованный текст."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        print(f"[ERROR] parse_markdown: Error reading {file_path} - {e}")
        return ""

    content = _strip_html_tags(content)
    lines = content.splitlines()
    parsed_lines = []

    for line in lines:
        stripped = line.strip()
        if stripped == "":
            parsed_lines.append("")
            continue

        # Заголовки (#, ##, ...)
        if re.match(r"#{1,6}\s+", stripped):
            header_text = re.sub(r"^#{1,6}\s+", "", stripped)
            parsed_lines.append(_normalize_links(header_text))
            continue

        # Маркированные списки (-, *, +)
        if re.match(r"(\*|\-|\+)\s+", stripped):
            bullet_text = re.sub(r"^(\*|\-|\+)\s+", "- ", stripped)
            parsed_lines.append(_normalize_links(bullet_text))
            continue

        # Нумерованные списки (1., 2., ...)
        if re.match(r"\d+\.\s+", stripped):
            numbered_text = re.sub(r"^\d+\.\s+", "1. ", stripped)
            parsed_lines.append(_normalize_links(numbered_text))
            continue

        # Таблицы с разделителями |
        if "|" in stripped:
            cells = [cell.strip() for cell in stripped.strip("|").split("|")]
            if all(cell == "" for cell in cells):
                continue
            if all(re.fullmatch(r"-+", cell) for cell in cells):
                continue
            parsed_lines.append(_normalize_links(" | ".join(cells)))
            continue

        parsed_lines.append(_normalize_links(stripped))

    result = "\n".join(parsed_lines).strip()
    print(f"[DEBUG] parse_markdown: Read {len(result)} chars from {file_path}")
    return result

def parse_document(file_path: str) -> str:
    """
    Определяет тип файла по расширению и вызывает соответствующий парсер.
    Поддерживает PDF, DOCX, XLS, XLSX и Markdown. Возвращает извлечённый
    текст или пустую строку, если парсинг не реализован или произошла ошибка.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return parse_pdf(file_path)
    elif ext == ".docx":
        return parse_docx(file_path)
    elif ext in [".xls", ".xlsx"]:
        return parse_excel(file_path)
    elif ext in [".md", ".markdown"]:
        return parse_markdown(file_path)
    else:
        print(f"[DEBUG] parse_document: Unsupported file extension {ext} for {file_path}")
        return ""
